# import all functions from the tkinter 
# without pywhatkit module; setup all functions in this script
   
import sys
import csv
import time
import os
import pathlib
import requests
import threading
import pyperclip
import pyautogui as pg
import webbrowser as web
from tkinter import *
from tkinter import filedialog
from PIL import Image
from platform import system
from urllib.parse import quote
from datetime import datetime
from typing import Optional
from urllib.parse import quote
from itertools import zip_longest
from openpyxl import load_workbook
pg.FAILSAFE = False
WIDTH, HEIGHT = pg.size()

class log:
    def format_message(self,message: str) -> str:
        """Formats the Message to remove redundant Spaces and Newline chars"""
        msg_l = message.split(" ")
        new = []
        for x in msg_l:
            if "\n" in x:
                x = x.replace("\n", "")
                new.append(x) if not len(x) == 0 else None
            elif len(x) != 0:
                new.append(x)
        return " ".join(new)


    def log_message(self,_time: time.struct_time, receiver: str, message: str) -> None:
        """Logs the Message Information after it is Sent"""

        if not os.path.exists("WpGui_DB.txt"):
            file = open("WpGui_DB.txt", "w+")
            file.close()

        message = self.format_message(message)

        with open("WpGui_DB.txt", "a", encoding="utf-8") as file:
            if w_core.check_number(receiver):
                file.write(
                    f"Date: {_time.tm_mday}/{_time.tm_mon}/{_time.tm_year}\nTime: {_time.tm_hour}:{_time.tm_min}\n"
                    f"Phone Number: {receiver}\nMessage: {message}"
                )
            else:
                file.write(
                    f"Date: {_time.tm_mday}/{_time.tm_mon}/{_time.tm_year}\nTime: {_time.tm_hour}:{_time.tm_min}\n"
                    f"Group ID: {receiver}\nMessage: {message}"
                )
            file.write("\n--------------------\n")
            file.close()


    def log_image(self,_time: time.struct_time, path: str, receiver: str, caption: str) -> None:
        """Logs the Image Information after it is Sent"""

        if not os.path.exists("WpGui_DB.txt"):
            file = open("WpGui_DB.txt", "w+")
            file.close()

        caption = self.format_message(caption)

        with open("WpGui_DB.txt", "a", encoding="utf-8") as file:
            if w_core.check_number(number=receiver):
                file.write(
                    f"Date: {_time.tm_mday}/{_time.tm_mon}/{_time.tm_year}\nTime: {_time.tm_hour}:{_time.tm_min}\n"
                    f"Phone Number: {receiver}\nImage: {path}\nCaption: {caption}"
                )

            else:
                file.write(
                    f"Date: {_time.tm_mday}/{_time.tm_mon}/{_time.tm_year}\nTime: {_time.tm_hour}:{_time.tm_min}\n"
                    f"Group ID: {receiver}\nImage: {path}\nCaption: {caption}"
                )
            file.write("\n--------------------\n")
            # file.close()

class core:
    def check_number(self,number: str) -> bool:
        """Checks the Number to see if contains the Country Code"""

        return number[1:].isdigit() #"+" in number or "_" in number


    def close_tab(self,wait_time: int = 2) -> None:
        """Closes the Currently Opened Browser Tab"""

        time.sleep(wait_time)
        if system().lower() in ("windows", "linux"):
            pg.hotkey("ctrl", "w")
        elif system().lower() == "darwin":
            pg.hotkey("command", "w")
        else:
            raise Warning(f"{system().lower()} not supported!")
        pg.press("enter")


    def check_connection(self) -> None:
        """Check the Internet connection of the Host Machine"""

        try:
            requests.get("https://google.com")
        except requests.RequestException:
            raise Exception(
                "Error while connecting to the Internet. Make sure you are connected to the Internet!"
            )


    def _web(self,receiver: str, message: str) -> None:
        """Opens WhatsApp Web based on the Receiver"""
        # if self.check_number(number=receiver):
        try:            
            web.open(
                "https://web.whatsapp.com/send?phone="
                + receiver)         # use receiver only and enter text later
                    
        except:
            raise Exception("Web browser couldn't load requested url with number")
            
        # else:
        #     pass
            # web.open("https://web.whatsapp.com/accept?code=" + receiver)  comment as Invalid url


    def send_message(self,message: str, receiver: str, wait_time: int) -> None:
        """Parses and Sends the Message"""

        self._web(receiver=receiver, message=message)
        time.sleep(wait_time) # WpGuiv1.3   [ time.sleep(7)]
        
        # pg.click(WIDTH / 2, HEIGHT / 2)
        # time.sleep(wait_time/2)  # WpGuiv1.3  [ time.sleep(wait_time -7)]
        # if not self.check_number(number=receiver):        # removed in v1.4

        # if test.count != 1:            
        #     for char in message:
        #         if char == "\n":
        #             pg.hotkey("shift", "enter")
        #         else:
        #             pg.typewrite(char)
        # pg.click(x=100, y=200)  # Adjust coordinates to your Text widget
        pg.hotkey('ctrl', 'v')
        try:            
            pg.press("enter")
            time.sleep(1)  
        except:
            pass
        finally:
            pg.press("enter")
        # time.sleep(1)


    def copy_image(self,path: str) -> None:
        """Copy the Image to Clipboard based on the Platform"""
        
        if system().lower() == "linux":
            if pathlib.Path(path).suffix in (".PNG", ".png"):
                os.system(f"copyq copy image/png - < {path}")
            elif pathlib.Path(path).suffix in (".jpg", ".JPG", ".jpeg", ".JPEG"):
                os.system(f"copyq copy image/jpeg - < {path}")
            else:
                raise Exception(
                    f"File Format {pathlib.Path(path).suffix} is not Supported!"
                )
        elif system().lower() == "windows":
            
            from io import BytesIO

            import win32clipboard
            # from PIL import Image            
            image = Image.open(path) #WpMsg.img_file_path            
            output = BytesIO()
            image.convert("RGB").save(output, "BMP")
            data = output.getvalue()[14:]
            output.close()            
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
            win32clipboard.CloseClipboard()

        elif system().lower() == "darwin":
            if pathlib.Path(path).suffix in (".jpg", ".jpeg", ".JPG", ".JPEG"):
                os.system(
                    f"osascript -e 'set the clipboard to (read (POSIX file \"{path}\") as JPEG picture)'"
                )
            else:
                raise Exception(
                    f"File Format {pathlib.Path(path).suffix} is not Supported!"
                )
        else:
            raise Exception(f"Unsupported System: {system().lower()}")     


    def send_image(self,path: str, caption: str, receiver: str, wait_time: int) -> None:
        """Sends the Image to a Contact or a Group based on the Receiver"""
        
        self.copy_image(path=path)
        time.sleep(3)
        self._web(message=caption, receiver=receiver)
        
        time.sleep(wait_time-3) # WpGuiv1.3        
        # pg.click(WIDTH / 2, HEIGHT / 2)       
        # self.copy_image(path=path)
        # if caption:                 
        #     for char in caption:
        #         if char == "\n":
        #             pg.hotkey("shift", "enter")
        #         else:
        #             pg.typewrite(char)
        
        if system().lower() == "darwin":
            pg.hotkey("command", "v")
        else:
            pg.hotkey("ctrl", "v")
        try:
            time.sleep(1)
            pg.press("enter")            
        except:
            pass
        finally:
            time.sleep(1)
            pg.press("enter")
        time.sleep(1)

class whatkit:
    def sendwhatmsg_instantly(self,
        phone_no: str,
        message: str,
        wait_time: int = 13,
        tab_close: bool = True,
        close_time: int = 3,
    ) -> None:
        """Send WhatsApp Message Instantly"""

        w_core.send_message(message=message, receiver=phone_no, wait_time=wait_time)
        w_log.log_message(_time=time.localtime(), receiver=phone_no, message=message)
        time.sleep(1)
        if tab_close:
            w_core.close_tab(wait_time=close_time)

    def sendwhatmsg(self,
        phone_no: str,
        message: str,
        time_hour: int,
        time_min: int,
        wait_time: int = 13,
        tab_close: bool = True,
        close_time: int = 3,
    ) -> None:
        """Send a WhatsApp Message at a Certain Time"""

        if time_hour not in range(25) or time_min not in range(60):
            raise Warning("Invalid Time Format!")

        current_time = time.localtime()
        left_time = datetime.strptime(
            f"{time_hour}:{time_min}:0", "%H:%M:%S"
        ) - datetime.strptime(
            f"{current_time.tm_hour}:{current_time.tm_min}:{current_time.tm_sec}",
            "%H:%M:%S",
        )

        if left_time.seconds < wait_time:
            raise Exception(
                "Call Time must be Greater than Wait Time as WhatsApp Web takes some Time to Load!"
            )

        sleep_time = left_time.seconds - wait_time
        print(
            f"In {sleep_time} Seconds WhatsApp will open and after {wait_time} Seconds Message will be Delivered!"
        )
        
        time.sleep(sleep_time)
        w_core.send_message(message=message, receiver=phone_no, wait_time=wait_time)
        w_log.log_message(_time=current_time, receiver=phone_no, message=message)
        time.sleep(1)
        if tab_close:
            w_core.close_tab(wait_time=close_time)

    def sendwhatmsg_to_group(self,
        group_id: str,
        message: str,
        time_hour: int,
        time_min: int,
        wait_time: int = 13,
        tab_close: bool = False,
        close_time: int = 3,
    ) -> None:
        """Send WhatsApp Message to a Group at a Certain Time"""

        if time_hour not in range(25) or time_min not in range(60):
            raise Warning("Invalid Time Format!")

        current_time = time.localtime()
        left_time = datetime.strptime(
            f"{time_hour}:{time_min}:0", "%H:%M:%S"
        ) - datetime.strptime(
            f"{current_time.tm_hour}:{current_time.tm_min}:{current_time.tm_sec}",
            "%H:%M:%S",
        )

        if left_time.seconds < wait_time:
            raise Exception(
                "Call Time must be Greater than Wait Time as WhatsApp Web takes some Time to Load!"
            )

        sleep_time = left_time.seconds - wait_time
        print(
            f"In {sleep_time} Seconds WhatsApp will open and after {wait_time} Seconds Message will be Delivered!"
        )
        
        time.sleep(sleep_time)
        w_core.send_message(message=message, receiver=group_id, wait_time=wait_time)
        w_log.log_message(_time=current_time, receiver=group_id, message=message)
        if tab_close:
            w_core.close_tab(wait_time=close_time)

    def sendwhatmsg_to_group_instantly(self,
        group_id: str,
        message: str,
        wait_time: int = 13,
        tab_close: bool = False,
        close_time: int = 3,
    ) -> None:
        """Send WhatsApp Message to a Group Instantly"""

        current_time = time.localtime()

        time.sleep(wait_time)
        w_core.send_message(message=message, receiver=group_id, wait_time=wait_time)
        w_log.log_message(_time=current_time, receiver=group_id, message=message)
        if tab_close:
            w_core.close_tab(wait_time=close_time)


    def sendwhats_image(self,
        receiver: str,
        img_path: str,
        caption: str = "",
        wait_time: int = 18,
        tab_close: bool = True,
        close_time: int = 3,
    ) -> None:
        """Send Image to a WhatsApp Contact or Group at a Certain Time"""        
        current_time = time.localtime()
        
        w_core.send_image(path=img_path, caption=caption, receiver=receiver, wait_time=wait_time)
        w_log.log_image(_time=current_time, path=img_path, receiver=receiver, caption=caption)
        if tab_close:
            w_core.close_tab(wait_time=close_time)

    def sendwhats_image_schedule(self,receiver: str,
        img_path: str,
        time_hour: int,
        time_min: int,
        caption: str = "",        
        wait_time: int = 18,
        tab_close: bool = True,
        close_time: int = 3,
    ) -> None:

        if time_hour not in range(25) or time_min not in range(60):
            raise Warning("Invalid Time Format!")

        current_time = time.localtime()
        left_time = datetime.strptime(
            f"{time_hour}:{time_min}:0", "%H:%M:%S"
        ) - datetime.strptime(
            f"{current_time.tm_hour}:{current_time.tm_min}:{current_time.tm_sec}",
            "%H:%M:%S",
        )

        if left_time.seconds < wait_time:
            raise Exception(
                "Call Time must be Greater than Wait Time as WhatsApp Web takes some Time to Load!"
            )

        sleep_time = left_time.seconds - wait_time
        print(
            f"In {sleep_time} Seconds WhatsApp will open and after {wait_time} Seconds Message will be Delivered!"
        )
        
        time.sleep(sleep_time)        
        w_core.send_image(path=img_path, caption=caption, receiver=receiver, wait_time=wait_time)
        w_log.log_image(_time=current_time, path=img_path, receiver=receiver, caption=caption)
        if tab_close:
            w_core.close_tab(wait_time=close_time)
        

class WpMsg:
    def __init__(self):        
        self.country_code = "+91"        
        self.wait_time = 15                 # for image preffered wait time 8
        self.close_time = 15                # for image preffered close time 10
        self.image =""
        self.num_list=[]
        self.message_list = []
        self.reciver_name_list = []
        # self.time_list = []
        self.message_text = ""  
        self.var = "Status" 
        self.csv_path = "No file selected..!"
        self.img_path = ""   
        self.img_file_path=""
        self.count = 1 
            
    def sendMessage(self) :          
        result = "Task Completed." 
        error = "ERROR: Sending Interupted..."  
        w_core.check_connection()        
        try:
            pg.click(self.TextArea.winfo_rootx() + 10, self.TextArea.winfo_rooty() + 10)  
            pg.hotkey('ctrl', 'a')  # Select all
            pg.hotkey('ctrl', 'c')  # Copy          
            # if self.message_list[1] != "":
            #     message = self.message_list[1]  
            # if self.message_list[1] != self.TextArea.get(1.0,END):
            message = self.TextArea.get(1.0,END)            
                  
        except IndexError:
            # self.var.set("WARNING: Select proper XLSX file")    
            message = self.TextArea.get(1.0,END)
        
        print("Message: {}".format(message))
        if message == "":
            self.var.set("WARNING: Not any Text message found")
            return False
        
        schedule_time = self.set_timeField.get() #self.time_list[1]
        if schedule_time != '':   
            hour = int(schedule_time[:2])
            minut = int(schedule_time[3:])        
            print("Scheduled Time: ",self.set_timeField.get())       

        if self.set_wait_timeField.get() != "":
            self.wait_time = int(self.set_wait_timeField.get())
        if self.set_close_timeField.get() != "":
            self.close_time = int(self.set_close_timeField.get())
        # self.count = 1               # added count in v1.4        
        
        wb = load_workbook(self.file_path)
        ws = wb.active  # Or wb['SheetName']
        
        for line in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):  
            if line[0] is None:                
                break           
            contact=str(line[0]).strip()
            pyperclip.copy(str(line[1]))            
            try:                    
                _message = message #if not self.message_list[index] else self.message_list[index]                
                if self.image == "y":                    
                    if not contact[1:].isdigit():                        
                        error = """ERROR:Not Possible Image in whatsapp-group."""                        
                        raise ValueError(error)                        
                    else:                        
                        try:
                            if self.count == 1 and schedule_time != '':
                                w_whats.sendwhats_image_schedule(contact, "{}".format(self.img_file_path), 
                                                        hour, minut, _message, 
                                                        self.wait_time, 
                                                        True, self.close_time)
                            else:                                                           
                                w_whats.sendwhats_image(contact, "{}".format(self.img_file_path), 
                                                                _message, self.wait_time, 
                                                                True, self.close_time)
                        except:
                            pass
                                        
                else:                                       
                    if not contact[1:].isdigit() :
                        print("contact: {}".format(contact))
                        try:
                            if self.count == 1 and schedule_time != '':              # added count in v1.4
                                w_whats.sendwhatmsg_to_group(contact, _message, 
                                                                hour, minut, 
                                                                self.wait_time, 
                                                                True, self.close_time)                    
                            else:
                                w_whats.sendwhatmsg_to_group_instantly(contact, 
                                                                    _message, 
                                                                    self.wait_time, 
                                                                    True, 
                                                                    self.close_time)   
                        except:                            
                            pass

                    else:
                        if self.count == 1 and schedule_time != '':                  # added count in v1.4
                            w_whats.sendwhatmsg(contact, _message, hour, 
                                                    minut , self.wait_time, 
                                                    True, self.close_time)          
                            
                        else:
                            w_whats.sendwhatmsg_instantly(contact, _message, 
                                                            self.wait_time, 
                                                            True, self.close_time)
                self.count +=1 
                time.sleep(2.5)        
            
        
            except KeyboardInterrupt:
                self.send_lable.config(fg='brown',font=("TimesNewRoman",12))
                self.var.set(str("Execution Terminate"))
                sys.exit(1)
            
            except Exception as e:
                self.send_lable.config(fg='red',font=("TimesNewRoman",12))
                self.var.set(error)
                # self.sendField.insert(10, str(error))
                print("{}\n{}".format(error,e))
        
        # self.sendField.insert(10, str(result)) 
        self.send_lable.config(fg='blue',font=("TimesNewRoman",12))
        self.var.set(result)
        print("+++++ Completed. +++++")         
        
    # Function for clearing the contents of all text entry boxes 
    def Clear(self): 

        self.num_list = []
        self.message_list = []
        self.reciver_name_list = []
        self.image = ""
        self.set_timeField.delete(0,'end')
        self.TextArea.delete(1.0,END)        
        self.var.set("Status")   
        self.csv_path.set("")
        self.img_path.set("")  
        self.set_wait_timeField.delete(0,'end')
        self.set_close_timeField.delete(0,'end')
        self.count=1

    # Function to open a specific .csv file
    def open_csv_file(self):        
        self.file_path = filedialog.askopenfilename(defaultextension=".xlsx", 
                                               filetypes=[("XLSX Files", "*.xlsx")])
        if self.file_path:
            self.csv_path.set(str(self.file_path.split("/")[-1]))  
            self.var.set("Status")       

    # Define the function to upload and save the image
    def upload_image(self):
        self.img_file_path = filedialog.askopenfilename()
        if self.img_file_path:            
            self.img_path.set(str(self.img_file_path.split("/")[-1]))
            try:
                image_file = Image.open(self.img_file_path) 
            except:
                print("WARNING:Image file not readable")
                pass           
            if image_file:
                self.image = "y"
                self.TextArea.insert(END,"Write Text Message Here..!")
                                
    def main(self):
        # Create a GUI window 
        gui = Tk()       
        gui.configure(background = "light grey")    
        gui.title("WhatsApp- Bulk Message Sender [SLS]")    
        gui.geometry("1000x600")   
        self.var = StringVar()
        self.csv_path = StringVar()
        self.img_path = StringVar()
        csv_text = Label(gui,text = " Browse XLSX File for numbers :",
                         bg = "Light Grey", font=("TimesNewRoman",12))                       
        message = Label(gui,text = " Message :",bg = "Light Grey",
                        font=("TimesNewRoman",12))                 
        set_time = Label(gui,text = "Enter Schedule Time If required :", 
                         bg = "Light Grey", font=("TimesNewRoman",12)) 
        img_text = Label(gui,text = " Do you Want to send Photo/Image? ",
                         bg = "Light Grey", font=("TimesNewRoman",12)) 
        self.send_lable =   Label(gui,textvariable=self.var, bg = "Dark Grey",
                                height=2, width = 63) 
        self.csv_status = Label(gui,textvariable=self.csv_path, bg = "Light Grey",
                                height=2, width = 20, font=("TimesNewRoman",12)) 
        self.img_status = Label(gui,textvariable=self.img_path, bg = "Light Grey",
                                height=2, width = 20, font=("TimesNewRoman",12)) 
        self.TextArea = Text(gui, height = 5, width = 55)
        # Create a button to open the .csv file
        open_button = Button(gui, text="Click Here to Upload XLSX file", 
                            command=self.open_csv_file)   
        upload_button = Button(gui, text="Click Here to Upload Image", 
                               command=self.upload_image)        
        clear = Button(gui, text = "   Clear   ", bg = "Dark Grey",
                       command = self.Clear)    #fg = "Black",  
        
        try:                  
            send = Button(gui, text = "   Send   ", bg = "Dark Grey",
                    command = lambda:threading.Thread(target=self.sendMessage).start())  
        except KeyboardInterrupt:
            sys.exit(1)
        
        set_wait_time = Label(gui,text = "Set Wait Time (Optional) :", 
                         bg = "Light Grey", font=("TimesNewRoman",12)) 
        set_close_time = Label(gui,text = "Set Close Time (Optional) :", 
                         bg = "Light Grey", font=("TimesNewRoman",12)) 
        

        csv_text.grid(row = 15, column = 20, padx = 10, pady = 20, ipady = 5)
        open_button.grid(row = 15, column = 25, padx = 10, pady = 20, ipady = 5) 
        self.csv_status.grid(row = 15, column = 26, ipady = 5, sticky='w') 
        img_text.grid(row = 35, column = 20, padx = 10, pady = 20, ipady = 5)     
        upload_button.grid(row = 35, column = 25, padx = 10, pady = 20, ipady = 5)
        self.img_status.grid(row = 35, column = 26, ipady = 5, sticky='w')  
        set_time.grid(row = 45, column = 20, padx = 10,  pady = 20, ipady = 5)
        message.grid(row = 55, column = 20, padx = 10, pady = 20, ipady = 5) 
        self.TextArea.grid(row = 55, column = 25, columnspan=5, sticky='w')          
        send.grid(row = 75, column = 20, padx = 10, pady = 20, ipady = 5) 
        clear.grid(row = 95, column = 20,padx = 10, pady = 20, ipady = 5)  
        set_wait_time.grid(row = 97, column = 20, padx = 10,  pady = 20, ipady = 5)  
        set_close_time.grid(row = 99, column = 20, padx = 10,  pady = 20, ipady = 5)    
        
        self.set_timeField = Entry(gui)        
        # self.sendField = Entry(gui)
        self.set_wait_timeField = Entry(gui)
        self.set_close_timeField = Entry(gui)            
        
        self.set_timeField.grid(row = 45, column = 25, ipady = 5, sticky='w')
        # self.sendField.grid(row = 75, column = 25, ipadx= 160, ipady = 20)  
        self.send_lable.grid(row = 75, column = 25, columnspan=2, sticky='w')

        self.set_wait_timeField.grid(row = 97, column = 25, ipady = 5, sticky='w')
        self.set_close_timeField.grid(row = 99, column = 25, ipady = 5, sticky='w')
        
        # Start the GUI 
        gui.mainloop() 

if __name__ == "__main__" :    
    test = WpMsg()
    w_core = core()
    w_whats = whatkit()
    w_log = log()
    test.main()
  
    